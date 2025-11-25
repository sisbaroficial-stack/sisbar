from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from inventario.models import Producto
from categorias.models import Categoria
from movimientos.models import Movimiento
from usuarios.views import registrar_actividad


@login_required
def reportes_home_view(request):
    """
    Página principal de reportes
    """
    # Estadísticas para mostrar en la página
    total_productos = Producto.objects.filter(activo=True).count()
    categorias_count = Categoria.objects.filter(activa=True).count()
    
    # Movimientos del último mes
    hace_30_dias = timezone.now() - timedelta(days=30)
    movimientos_mes = Movimiento.objects.filter(fecha__gte=hace_30_dias).count()
    
    context = {
        'total_productos': total_productos,
        'categorias_count': categorias_count,
        'movimientos_mes': movimientos_mes,
    }
    
    return render(request, 'reportes/home.html', context)


@login_required
def exportar_productos_excel(request):
    """
    Exportar lista de productos a Excel
    """
    # Crear libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Estilos
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:I1')
    titulo = ws['A1']
    titulo.value = "REPORTE DE INVENTARIO - SISBAR "
    titulo.font = Font(bold=True, size=16, color="667EEA")
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    
    # Fecha de generación
    ws.merge_cells('A2:I2')
    fecha = ws['A2']
    fecha.value = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    fecha.alignment = Alignment(horizontal='center')
    
    # Espacio
    ws.append([])
    
    # Encabezados
    headers = ['Código', 'Nombre', 'Categoría', 'Subcategoría', 'Cantidad', 
               'Unidad', 'Estado', 'Precio', 'Proveedor']
    ws.append(headers)
    
    # Estilo de encabezados
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    productos = Producto.objects.filter(activo=True).select_related(
        'categoria', 'subcategoria', 'proveedor'
    )
    
    for producto in productos:
        ws.append([
            producto.codigo,
            producto.nombre,
            producto.categoria.nombre,
            producto.subcategoria.nombre if producto.subcategoria else 'N/A',
            producto.cantidad,
            producto.get_unidad_medida_display(),
            producto.get_estado_display(),
            float(producto.precio_compra),
            producto.proveedor.nombre if producto.proveedor else 'N/A'
        ])
    
    # Aplicar bordes a todas las celdas de datos
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = border
            if cell.column == 5:  # Cantidad
                cell.alignment = Alignment(horizontal='center')
            if cell.column == 8:  # Precio
                cell.number_format = '$#,##0.00'
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 25
    
    # Registrar actividad
    registrar_actividad(
        request.user,
        'EXPORTAR',
        'Exportó reporte de inventario a Excel',
        request
    )
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=inventario_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


@login_required
def exportar_productos_pdf(request):
    """
    Exportar lista de productos a PDF
    """
    # Crear documento
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=inventario_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    # Crear PDF
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#667EEA'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Título
    title = Paragraph("REPORTE DE INVENTARIO", title_style)
    elements.append(title)
    
    subtitle = Paragraph(
        f"SISBAR  - Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal']
    )
    elements.append(subtitle)
    elements.append(Spacer(1, 20))
    
    # Datos de la tabla
    data = [['Código', 'Nombre', 'Categoría', 'Cantidad', 'Estado', 'Precio']]
    
    productos = Producto.objects.filter(activo=True).select_related('categoria')
    
    for producto in productos:
        data.append([
            producto.codigo,
            producto.nombre[:30],  # Truncar nombre largo
            producto.categoria.nombre,
            f"{producto.cantidad} {producto.get_unidad_medida_display()}",
            producto.get_estado_display(),
            f"${producto.precio_compra:,.2f}"
        ])
    
    # Crear tabla
    table = Table(data, colWidths=[1.2*inch, 2.5*inch, 1.5*inch, 1.2*inch, 1.2*inch, 1.2*inch])
    
    # Estilo de la tabla
    table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Contenido
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),
        ('ALIGN', (4, 1), (4, -1), 'CENTER'),
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        
        # Bordes
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BOX', (0, 0), (-1, -1), 2, colors.black),
    ]))
    
    elements.append(table)
    
    # Pie de página con estadísticas
    elements.append(Spacer(1, 20))
    
    stats_text = f"""
    <b>Estadísticas:</b><br/>
    Total de productos: {productos.count()}<br/>
    Productos disponibles: {productos.filter(estado='DISPONIBLE').count()}<br/>
    Productos por agotarse: {productos.filter(estado='POR_AGOTAR').count()}<br/>
    Productos agotados: {productos.filter(estado='AGOTADO').count()}
    """
    
    stats = Paragraph(stats_text, styles['Normal'])
    elements.append(stats)
    
    # Registrar actividad
    registrar_actividad(
        request.user,
        'EXPORTAR',
        'Exportó reporte de inventario a PDF',
        request
    )
    
    # Construir PDF
    doc.build(elements)
    
    return response


@login_required
def exportar_movimientos_excel(request):
    """
    Exportar movimientos a Excel
    """
    # Obtener rango de fechas
    dias = int(request.GET.get('dias', 30))
    fecha_desde = timezone.now() - timedelta(days=dias)
    
    # Crear libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    
    # Estilos
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Título
    ws.merge_cells('A1:G1')
    titulo = ws['A1']
    titulo.value = f"REPORTE DE MOVIMIENTOS - Últimos {dias} días"
    titulo.font = Font(bold=True, size=16, color="667EEA")
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    
    # Encabezados
    ws.append([])
    ws.append([])
    headers = ['Fecha', 'Tipo', 'Producto', 'Código', 'Cantidad', 'Usuario', 'Motivo']
    ws.append(headers)
    
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    movimientos = Movimiento.objects.filter(
        fecha__gte=fecha_desde
    ).select_related('producto', 'usuario').order_by('-fecha')
    
    for mov in movimientos:
        ws.append([
            mov.fecha.strftime('%d/%m/%Y %H:%M'),
            mov.get_tipo_display(),
            mov.producto.nombre,
            mov.producto.codigo,
            mov.cantidad,
            mov.usuario.username if mov.usuario else "Sistema",
            mov.motivo or 'N/A'
        ])
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30
    
    # Registrar actividad
    registrar_actividad(
        request.user,
        'EXPORTAR',
        f'Exportó reporte de movimientos ({dias} días) a Excel',
        request
    )
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=movimientos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response